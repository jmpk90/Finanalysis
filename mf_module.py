"""
EQUITEX PRO — Mutual Fund & CAS Portfolio Module
_VERSION = "v6-clean-errors"  # update check: if this prints, correct file is loaded
==================================================
• Multi-account: import CAS statements for self + spouse + family
• casparser library for CDSL / NSDL / CAMS / KFintech PDF parsing
• pdfplumber fallback regex parser if casparser not installed
• Live NAV via mfapi.in (AMFI, free, no key needed)
• Fund search by name → auto-link AMFI scheme code
• Per-account and combined portfolio views
• Returns analysis: absolute, CAGR, XIRR estimate, period NAVs
• Auto-syncs totals → Finance Advisor
"""

import streamlit as st
import json, os, re, io, datetime, tempfile
import requests

# ── Constants ─────────────────────────────────────────────────
MFAPI_SEARCH = "https://api.mfapi.in/mf/search?q={}"
MFAPI_LATEST = "https://api.mfapi.in/mf/{}/latest"
MFAPI_HIST   = "https://api.mfapi.in/mf/{}"
CACHE_TTL    = 3600

MF_CATEGORIES = [
    "Large Cap","Mid Cap","Small Cap","Flexi Cap","Multi Cap",
    "ELSS (Tax Saver)","Index Fund","ETF","Debt / Liquid",
    "Hybrid / Balanced","International","Sectoral / Thematic","Other"
]

ACCOUNT_RELATIONS = ["Self","Spouse","Father","Mother","Son","Daughter","Other"]

# ── Disk persistence ──────────────────────────────────────────
def get_store() -> dict:
    from equitex_store import get_mf_store
    store = get_mf_store()
    st.session_state.mf_store = store
    return store

def save_store(store: dict):
    st.session_state.mf_store = store
    from equitex_store import save_mf_store
    save_mf_store(store)

# ── Account helpers ───────────────────────────────────────────
def get_accounts(store: dict) -> dict:
    return store.setdefault("accounts", {})

def all_funds(store: dict) -> list:
    out = []
    for acc_id, acc in store.get("accounts", {}).items():
        for f in acc.get("funds", []):
            out.append({**f, "_account_id": acc_id,
                        "_account_name": acc.get("name", acc_id),
                        "_relation": acc.get("relation", "")})
    for f in store.get("funds", []):
        out.append({**f, "_account_id": "manual",
                    "_account_name": "Manual", "_relation": "Self"})
    return out

# ── mfapi.in helpers ──────────────────────────────────────────
@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def search_funds(q: str) -> list:
    if len(q) < 3: return []
    try:
        r = requests.get(MFAPI_SEARCH.format(requests.utils.quote(q)), timeout=8)
        return r.json()[:20] if r.status_code == 200 else []
    except Exception:
        return []

@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def fetch_latest_nav(code) -> dict:
    try:
        r = requests.get(MFAPI_LATEST.format(code), timeout=8)
        if r.status_code == 200:
            d = r.json().get("data", [])
            if d:
                return {"nav": float(d[0]["nav"]), "date": d[0]["date"]}
    except Exception:
        pass
    return {}

@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def fetch_nav_history(code, days=365) -> list:
    try:
        r = requests.get(MFAPI_HIST.format(code), timeout=12)
        return r.json().get("data", [])[:days] if r.status_code == 200 else []
    except Exception:
        return []

# ── Financial calcs ───────────────────────────────────────────
def abs_ret(inv, cur):
    return (cur - inv) / inv * 100 if inv > 0 else 0.0

def cagr(inv, cur, years):
    if inv <= 0 or cur <= 0 or years <= 0: return 0.0
    return ((cur / inv) ** (1.0 / years) - 1) * 100

def xirr_est(inv, cur, sip, since_iso):
    try:
        since = datetime.date.fromisoformat(since_iso)
        years = max(0.08, (datetime.date.today() - since).days / 365.25)
    except Exception:
        years = 1.0
    if sip > 0:
        return cagr(inv + sip * years * 12, cur, years / 2)
    return cagr(inv, cur, years)

def nav_period_ret(history, days):
    if len(history) < days: return None
    try:
        latest = float(history[0]["nav"])
        past   = float(history[days - 1]["nav"])
        return (latest - past) / past * 100
    except Exception:
        return None

def fmt(n):
    if n is None: return "—"
    n = float(n)
    if n >= 1e7:  return f"Rs {n/1e7:.2f} Cr"
    if n >= 1e5:  return f"Rs {n/1e5:.2f} L"
    if n >= 1e3:  return f"Rs {n/1e3:.1f}K"
    return f"Rs {n:,.0f}"

def fmt_inr(n): return fmt(n)

def mask_pan(pan):
    """Show only the last 4 characters of a PAN — the rest is masked
    since PAN is a sensitive personal identifier, not something that
    needs to be fully visible on screen for day-to-day use."""
    pan = str(pan or "").strip()
    if len(pan) < 5:
        return pan if pan else "N/A"
    return "•" * (len(pan) - 4) + pan[-4:]

def color_ret(v):
    return "#3ecf8e" if v >= 0 else "#e05252"

# ══════════════════════════════════════════════════════════════
# CAS PARSING
# ══════════════════════════════════════════════════════════════


def parse_mf_excel(file_bytes: bytes) -> dict:
    """Parse EQUITEX PRO Excel template — reads Portfolio Summary sheet."""
    try:
        import openpyxl, io as _io, re as _re
        wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), data_only=True)

        def _sheet(*names):
            low = {s.lower().replace(" ",""): s for s in wb.sheetnames}
            for n in names:
                k = n.lower().replace(" ","")
                if k in low: return wb[low[k]]
            return None

        def _v(c): return str(c.value).strip() if c.value is not None else ""
        def _n(c):
            v = c.value
            if v is None: return 0.0
            if isinstance(v,(int,float)): return float(v)
            try: return float(str(v).replace(",","").replace("₹","").strip())
            except: return 0.0

        # ── Accounts ──────────────────────────────────
        accs = {}
        ws_a = _sheet("accounts")
        if ws_a:
            for row in ws_a.iter_rows(min_row=3):
                cs = list(row)
                if not cs or cs[0].value is None: continue
                aid = _v(cs[0])
                if not aid or aid.lower()=="account id": continue
                accs[aid] = {
                    "name":     _v(cs[1]) if len(cs)>1 else "",
                    "relation": _v(cs[2]) if len(cs)>2 else "Self",
                    "pan":      _v(cs[3]) if len(cs)>3 else "",
                }

        # ── Portfolio Summary (primary) ────────────────
        by_acc = {}
        ws_s = _sheet("portfoliosummary","portfolio summary","portfolio")
        if ws_s:
            hrow = None
            for r in ws_s.iter_rows(min_row=1, max_row=8):
                for c in r:
                    if str(c.value or "").strip().lower() in ("account id","closing units","isin"):
                        hrow = c.row; break
                if hrow: break
            if hrow:
                col = {}
                for c in ws_s[hrow]:
                    h = _re.sub(r'\s+',' ', str(c.value or "").lower().replace("(","").replace(")","").strip())
                    if "account id"    in h: col["aid"]      = c.column
                    if "fund name"     in h: col["name"]     = c.column
                    if "isin"          in h: col["isin"]     = c.column
                    if "folio"         in h: col["folio"]    = c.column
                    if "closing units" in h: col["units"]    = c.column
                    if "invested"      in h: col["invested"] = c.column
                    if "fund house"    in h: col["amc"]      = c.column
                for row in ws_s.iter_rows(min_row=hrow+1):
                    cs = list(row)
                    if not cs: continue
                    aid  = _v(cs[col["aid"]-1])  if "aid"  in col else ""
                    name = _v(cs[col["name"]-1]) if "name" in col else ""
                    if not aid or not name: continue
                    u = _n(cs[col["units"]-1])    if "units"    in col else 0.0
                    i = _n(cs[col["invested"]-1]) if "invested" in col else 0.0
                    if u <= 0: continue
                    by_acc.setdefault(aid, []).append({
                        "scheme": name[:80],
                        "isin":   _v(cs[col["isin"]-1])  if "isin"  in col else "",
                        "folio":  _v(cs[col["folio"]-1]) if "folio" in col else "",
                        "amc":    _v(cs[col["amc"]-1])   if "amc"   in col else "",
                        "amfi":"", "units":round(u,3), "nav":0.0,
                        "value":0.0, "cost":i, "valuation_date":"",
                        "_account_id": aid,
                    })

        # ── Transactions fallback ──────────────────────
        if not by_acc:
            ws_t = _sheet("transactions")
            if ws_t:
                hrow=None
                for r in ws_t.iter_rows(min_row=1,max_row=8):
                    for c in r:
                        if str(c.value or "").strip().lower() in ("account id","transaction type"):
                            hrow=c.row; break
                    if hrow: break
                if hrow:
                    hdr={}
                    for c in ws_t[hrow]:
                        h=_re.sub(r'\s+',' ',str(c.value or "").lower().replace("(","").replace(")","").strip())
                        hdr[h]=c.column
                    COL={}
                    for canon,variants in {
                        "aid":   ["account id"],
                        "house": ["fund house"],
                        "name":  ["fund name & plan","fund name","scheme"],
                        "isin":  ["isin"],"folio":["folio / ucc","folio"],
                        "type":  ["transaction type"],
                        "amt":   ["amount ₹","amount rs","amount"],
                        "nav":   ["nav ₹","nav rs","nav"],
                        "units": ["units"],
                    }.items():
                        for v in variants:
                            if v in hdr: COL[canon]=hdr[v]; break
                    PURCH={"sip purchase","lumpsum purchase","switch in","dividend reinvestment","opening balance"}
                    ledger={}
                    for row in ws_t.iter_rows(min_row=hrow+1):
                        cs=list(row)
                        if not cs: continue
                        gv=lambda k: _v(cs[COL[k]-1]) if k in COL else ""
                        gn=lambda k: _n(cs[COL[k]-1]) if k in COL else 0.0
                        aid=gv("aid"); nm=gv("name")
                        if not aid or not nm: continue
                        isin=gv("isin"); tx=gv("type").lower()
                        uv=gn("units"); nv=gn("nav"); av=gn("amt")
                        fk=(aid,isin or nm)
                        if fk not in ledger:
                            ledger[fk]={"aid":aid,"name":nm,"isin":isin,"folio":gv("folio"),
                                "house":gv("house"),"open":0.0,"close":None,
                                "pu":0.0,"ru":0.0,"inv":0.0,"nav":0.0}
                        rd=ledger[fk]
                        if nv>0: rd["nav"]=nv
                        if tx=="opening balance": rd["open"]=uv
                        elif tx=="closing balance": rd["close"]=uv
                        elif tx in PURCH: rd["pu"]+=uv; rd["inv"]+=av if av>0 else uv*nv
                        elif tx in ("redemption","switch out"): rd["ru"]+=uv
                    for fk,rd in ledger.items():
                        cl=rd["close"] if rd["close"] is not None else rd["open"]+rd["pu"]-rd["ru"]
                        if cl<=0: continue
                        cost=rd["inv"] if rd["inv"]>0 else round(cl*rd["nav"],2)
                        by_acc.setdefault(rd["aid"],[]).append({
                            "scheme":rd["name"][:80],"isin":rd["isin"],"amfi":"",
                            "folio":rd["folio"],"amc":rd["house"],
                            "units":round(cl,3),"nav":rd["nav"],
                            "value":round(cl*rd["nav"],2) if rd["nav"]>0 else 0.0,
                            "cost":cost,"valuation_date":"","_account_id":rd["aid"],
                        })

        if not by_acc:
            return {"ok":False,"errors":["No fund data found. Fill Portfolio Summary sheet."]}

        folios=[]
        for aid,schemes in by_acc.items():
            ai=accs.get(aid,{})
            folios.append({"folio":aid,"amc":"Multiple","pan":ai.get("pan",""),
                "_account_id":aid,"_account_name":ai.get("name",aid),
                "_relation":ai.get("relation","Self"),"schemes":schemes})

        fa=next(iter(accs.values()),{})
        tf=sum(len(x["schemes"]) for x in folios)
        return {"ok":True,"source":"excel_template","file_type":"EXCEL",
            "investor":{"name":fa.get("name",""),"pan":fa.get("pan",""),"email":"","mobile":""},
            "statement_period":{"from":"","to":""},"folios":folios,
            "errors":[f"ℹ️ {len(folios)} account(s), {tf} fund(s) parsed from Excel"]}

    except ImportError:
        return {"ok":False,"errors":["openpyxl not installed: pip install openpyxl"]}
    except Exception as e:
        import traceback
        return {"ok":False,"errors":[f"Excel error: {e} | {traceback.format_exc().splitlines()[-1]}"]}

def parse_cas_pdf(file_bytes: bytes, password: str = "") -> dict:
    """
    Parse CDSL/NSDL CAS PDF.
    Tries 5 extraction methods in order — one will always work.
    """
    errors  = []
    raw_text = None

    # ── helper: try all pdfplumber modes without crashing on color PDFs ──
    def _try_pdfplumber(source):
        import pdfplumber
        open_kw = {}
        if password and password.strip():
            open_kw["password"] = password.strip()
        pages = []
        with pdfplumber.open(source, **open_kw) as pdf:
            for pg in pdf.pages:
                t = None
                # Mode 1: extract_words (bypasses DeviceRGB/Decimal crash)
                try:
                    words = pg.extract_words(x_tolerance=3, y_tolerance=3,
                                             keep_blank_chars=False, use_text_flow=True)
                    if words:
                        from itertools import groupby
                        ws = sorted(words, key=lambda w: (round(w["top"] / 5) * 5, w["x0"]))
                        lines = [" ".join(w["text"] for w in g)
                                 for _, g in groupby(ws, key=lambda w: round(w["top"] / 5) * 5)]
                        t = "\n".join(lines)
                except Exception:
                    pass
                # Mode 2: extract_text_simple
                if not t:
                    try:
                        t = pg.extract_text_simple()
                    except Exception:
                        pass
                # Mode 3: plain extract_text
                if not t:
                    try:
                        t = pg.extract_text()
                    except Exception:
                        pass
                if t:
                    pages.append(t)
        return "\n".join(pages)

    # ── 1. pdfminer — most robust, handles CDSL color PDFs natively ──────
    try:
        from pdfminer.high_level import extract_text as _pdfminer_extract
        from pdfminer.layout import LAParams
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                tmp.write(file_bytes)
                tmp_path = tmp.name
            raw_text = _pdfminer_extract(tmp_path, laparams=LAParams())
        finally:
            if tmp_path:
                try: os.unlink(tmp_path)
                except Exception: pass
        if raw_text and raw_text.strip():
            errors.append("ℹ️ Parsed via pdfminer")
    except ImportError:
        pass
    except Exception as e:
        errors.append(f"pdfminer error: {e}")

    # ── 2. pypdf ──────────────────────────────────────────────────────────
    if not (raw_text and raw_text.strip()):
        try:
            import pypdf
            reader = pypdf.PdfReader(io.BytesIO(file_bytes))
            if reader.is_encrypted and password:
                reader.decrypt(password)
            pages = []
            for pg in reader.pages:
                t = pg.extract_text()
                if t: pages.append(t)
            raw_text = "\n".join(pages)
            if raw_text and raw_text.strip():
                errors.append("ℹ️ Parsed via pypdf")
        except ImportError:
            pass
        except Exception as e:
            errors.append(f"pypdf error: {e}")

    # ── 3. pdfplumber (temp file) ─────────────────────────────────────────
    if not (raw_text and raw_text.strip()):
        try:
            import pdfplumber
            tmp_path = None
            try:
                with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                    tmp.write(file_bytes)
                    tmp_path = tmp.name
                raw_text = _try_pdfplumber(tmp_path)
                if raw_text and raw_text.strip():
                    errors.append("ℹ️ Parsed via pdfplumber")
            finally:
                if tmp_path:
                    try: os.unlink(tmp_path)
                    except Exception: pass
        except ImportError:
            errors.append("pdfplumber not installed — run: pip install pdfplumber")
        except Exception as e:
            errors.append(f"pdfplumber error: {e}")

    # ── 4. pymupdf (fitz) ─────────────────────────────────────────────────
    if not (raw_text and raw_text.strip()):
        try:
            import fitz
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            if doc.is_encrypted and password:
                doc.authenticate(password)
            pages = [pg.get_text("text") for pg in doc]
            doc.close()
            raw_text = "\n".join(pages)
            if raw_text and raw_text.strip():
                errors.append("ℹ️ Parsed via pymupdf")
        except ImportError:
            pass
        except Exception as e:
            errors.append(f"pymupdf error: {e}")

    # ── 5. casparser (last — known to fail on this CDSL format) ──────────
    if not (raw_text and raw_text.strip()):
        try:
            import casparser as _cas
            tmp_path = None
            try:
                with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                    tmp.write(file_bytes)
                    tmp_path = tmp.name
                raw = _cas.read_cas_pdf(tmp_path, password)
            finally:
                if tmp_path:
                    try: os.unlink(tmp_path)
                    except Exception: pass

            investor = {
                "name":   str(getattr(raw.investor_info, "name",   "") or ""),
                "email":  str(getattr(raw.investor_info, "email",  "") or ""),
                "pan":    "",
                "mobile": str(getattr(raw.investor_info, "mobile", "") or ""),
            }
            folios = []
            for folio in (raw.folios or []):
                schemes = []
                for s in (folio.schemes or []):
                    val = s.valuation or {}
                    schemes.append({
                        "scheme":         str(s.scheme or ""),
                        "isin":           str(getattr(s, "isin",  "") or ""),
                        "amfi":           str(getattr(s, "amfi",  "") or ""),
                        "units":          float(getattr(s, "close", 0) or 0),
                        "nav":            float(getattr(val, "nav",   0) or 0),
                        "value":          float(getattr(val, "value", 0) or 0),
                        "cost":           float(getattr(val, "cost",  0) or 0),
                        "valuation_date": str(getattr(val, "date",  "") or ""),
                    })
                folios.append({
                    "folio":   str(getattr(folio, "folio", "") or ""),
                    "amc":     str(getattr(folio, "amc",   "") or ""),
                    "pan":     str(getattr(folio, "PAN",   "") or ""),
                    "schemes": schemes,
                })
            return {
                "ok": True, "source": "casparser",
                "investor": investor,
                "statement_period": {
                    "from": str(getattr(raw.statement_period, "from", "") or ""),
                    "to":   str(getattr(raw.statement_period, "to",   "") or ""),
                },
                "file_type": str(getattr(raw, "file_type", "CAS") or "CAS"),
                "folios": folios, "errors": [],
            }
        except ImportError:
            pass
        except Exception as e:
            errors.append(f"casparser error: {e}")

    # ── Parse whatever text we got ────────────────────────────────────────
    if raw_text and raw_text.strip():
        result = _parse_cdsl_text(raw_text)
        # Remove info-only messages from errors shown to user
        user_errors = [e for e in errors if not e.startswith("ℹ️")]
        result["errors"] = user_errors + result.get("errors", [])
        return result

    return {
        "ok": False, "source": "error",
        "investor": {}, "folios": [],
        "statement_period": {}, "file_type": "",
        "errors": [e for e in errors if not e.startswith("ℹ️")]
                  or ["Could not extract text from PDF. Install pypdf: pip install pypdf"],
    }


def _parse_cdsl_text(text: str) -> dict:
    """
    Dedicated CDSL CAS PDF parser.

    The CDSL PDF has two useful sections:
      A) Account Details — lists AMC Name, Scheme Name, ISIN, Folio for each fund
      B) Holdings Summary Table — tabular rows with ISIN, folio, units, NAV, invested, value

    We build the scheme name & AMC from section A, and units/NAV/values from section B.
    The ISIN is the reliable join key between both sections.
    """
    # Strip non-ASCII (handles Hindi/Devanagari overlay text in CDSL PDFs)
    clean = re.sub(r'[^\x00-\x7F]+', ' ', text)
    clean = re.sub(r'[ \t]+', ' ', clean)

    errors = []

    # ── Investor info ─────────────────────────────────────────
    investor = {"name": "", "email": "", "pan": "", "mobile": ""}
    m = re.search(r'([A-Z][A-Z ]{3,40})\s*\(\s*PAN\s*[:\s]+([A-Z]{5}\d{4}[A-Z])\s*\)', clean)
    if m:
        investor["name"] = m.group(1).strip()
        investor["pan"]  = m.group(2).strip()
    else:
        m = re.search(r'PAN\s*[:\s]+([A-Z]{5}\d{4}[A-Z])', clean)
        if m: investor["pan"] = m.group(1).strip()
    m = re.search(r'Email\s*[Ii]d?\s*[:\s]+([\w.\-+]+@[\w.\-]+)', clean)
    if m: investor["email"] = m.group(1).strip()
    m = re.search(r'Mobile\s*No\s*[:\s]+(\d[\dX]{9})', clean, re.I)
    if m: investor["mobile"] = m.group(1).strip()

    # ── Statement period ──────────────────────────────────────
    period = {"from": "", "to": ""}
    m = re.search(
        r'(?:from|FROM)\s+(\d{2}[/-][A-Za-z]{3}[/-]\d{4}|\d{2}-\d{2}-\d{4})'
        r'.{0,30}?(?:to|TO)\s+(\d{2}[/-][A-Za-z]{3}[/-]\d{4}|\d{2}-\d{2}-\d{4})',
        clean, re.I
    )
    if m: period = {"from": m.group(1), "to": m.group(2)}

    file_type = "CDSL" if "CDSL" in clean else ("NSDL" if "NSDL" in clean else "CAS")

    # ── Build ISIN → scheme name lookup from Account Details ──
    # Pattern: "Scheme Name : <Name> Scheme Code : <CODE>"
    # then "ISIN : INF..."
    isin_to_scheme = {}
    isin_to_amc    = {}
    isin_to_folio  = {}

    for m in re.finditer(r'AMC Name\s*:\s*([^\n]+)', clean):
        amc_name = m.group(1).strip()
        # Scan forward up to 600 chars for Scheme Name and ISIN
        ctx = clean[m.start(): m.start()+600]
        sm = re.search(r'Scheme Name\s*:\s*(.*?)(?:Scheme Code|ISIN)', ctx, re.S)
        im = re.search(r'ISIN\s*:\s*(IN[FE][A-Z0-9]{9})', ctx)
        fm = re.search(r'Folio\s*No\s*:\s*([\w/]+)', ctx)
        if im:
            isin = im.group(1)
            if isin not in isin_to_amc:
                isin_to_amc[isin] = amc_name
            if sm and isin not in isin_to_scheme:
                scheme_raw = re.sub(r'\s+', ' ', sm.group(1)).strip()
                isin_to_scheme[isin] = scheme_raw
            if fm and isin not in isin_to_folio:
                isin_to_folio[isin] = fm.group(1)

    # ── Parse holdings table rows ─────────────────────────────
    # Row format (one line):  ISIN  folio  units  nav  invested  valuation  pl  pl%
    ISIN_ROW = re.compile(
        r'\b(IN[FE][A-Z0-9]{9})\b\s+'
        r'([\d]+(?:/[\d]+)?)\s+'   # folio number
        r'([\d,]+\.\d+)\s+'        # closing balance (units) — must have decimal
        r'([\d,]+\.?\d*)\s+'       # NAV
        r'([\d,]+\.?\d*)\s+'       # cumulative invested
        r'([\d,]+\.?\d*)\s+'       # valuation
        r'(-?[\d,]+\.?\d*)\s+'     # unrealised P/L
        r'(-?[\d,]+\.?\d*)'        # unrealised P/L %
    )

    funds_raw = []
    seen = set()
    for m in ISIN_ROW.finditer(clean):
        isin  = m.group(1)
        folio = m.group(2)
        key   = (isin, folio)
        if key in seen: continue
        seen.add(key)

        try:
            units    = float(m.group(3).replace(',', ''))
            nav      = float(m.group(4).replace(',', ''))
            invested = float(m.group(5).replace(',', ''))
            value    = float(m.group(6).replace(',', ''))
        except Exception:
            continue

        if units <= 0 or value <= 0:
            continue

        # Scheme name: prefer account-details lookup, fall back to surrounding text
        scheme = isin_to_scheme.get(isin, '')
        if not scheme:
            pre = clean[max(0, m.start()-300): m.start()]
            pre_lines = [l.strip() for l in pre.split('\n')
                         if l.strip() and len(l.strip()) > 4
                         and not re.match(r'^[\d\s,.%\-]+$', l.strip())
                         and not re.match(r'^(Page |INR|Profit|Cumul|Closing|Opening|Stamp|ISIN|Grand|Scheme)', l.strip(), re.I)]
            if pre_lines:
                scheme = re.sub(r'^[A-Z0-9]{2,6}\s*-\s*', '', ' '.join(pre_lines[-3:])).strip()

        funds_raw.append({
            "scheme":         (scheme or isin)[:80],
            "isin":           isin,
            "amfi":           "",
            "units":          units,
            "nav":            nav,
            "value":          value,
            "cost":           invested,
            "valuation_date": period.get("to", ""),
            "_amc":           isin_to_amc.get(isin, ""),
            "_folio":         folio,
        })

    if not funds_raw:
        errors.append("No fund rows found in holdings table. "
                      "The PDF may need pymupdf for better text extraction: pip install pymupdf")

    # ── Group into folios by folio number ────────────────────
    folio_groups: dict = {}
    for s in funds_raw:
        fkey = s.get("_folio", "unknown")
        if fkey not in folio_groups:
            folio_groups[fkey] = {
                "folio":   fkey,
                "amc":     s.get("_amc", ""),
                "pan":     investor.get("pan", ""),
                "schemes": []
            }
        clean_s = {k: v for k, v in s.items() if not k.startswith("_")}
        folio_groups[fkey]["schemes"].append(clean_s)

    folios = list(folio_groups.values())

    return {
        "ok": bool(funds_raw),
        "investor": investor,
        "statement_period": period,
        "file_type": file_type,
        "folios": folios,
        "errors": errors,
    }


def _regex_parse_cas(text: str) -> dict:
    """Generic fallback for non-CDSL formats."""
    clean = re.sub(r'[^\x00-\x7F]+', ' ', text)
    investor = {"name": "", "email": "", "pan": "", "mobile": ""}
    m = re.search(r'([A-Z][A-Z ]{3,40})\s*\(\s*PAN\s*[:\s]+([A-Z]{5}\d{4}[A-Z])\s*\)', clean)
    if m:
        investor["name"] = m.group(1).strip()
        investor["pan"]  = m.group(2).strip()
    period = {"from": "", "to": ""}
    file_type = "CDSL" if "CDSL" in clean else ("NSDL" if "NSDL" in clean else "CAS")
    ISIN_ROW = re.compile(
        r'\b(IN[FE][A-Z0-9]{9})\b\s+([\d]+(?:/[\d]+)?)\s+([\d,]+\.\d+)\s+'
        r'([\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+(-?[\d,]+\.?\d*)\s+(-?[\d,]+\.?\d*)'
    )
    schemes, seen = [], set()
    for m in ISIN_ROW.finditer(clean):
        key = (m.group(1), m.group(2))
        if key in seen: continue
        seen.add(key)
        try:
            schemes.append({
                "scheme": m.group(1), "isin": m.group(1), "amfi": "",
                "units": float(m.group(3).replace(',','')),
                "nav":   float(m.group(4).replace(',','')),
                "value": float(m.group(6).replace(',','')),
                "cost":  float(m.group(5).replace(',','')),
                "valuation_date": "",
            })
        except Exception: pass
    folios = [{"folio":"","amc":"","pan":"","schemes":schemes}] if schemes else []
    return {"ok": bool(schemes), "investor": investor, "statement_period": period,
            "file_type": file_type, "folios": folios, "errors": []}




# ── Convert parsed CAS → fund dicts ──────────────────────────
def cas_to_funds(parsed: dict) -> list:
    """Convert parsed data to fund objects — and immediately look up each
    fund's LIVE NAV by ISIN/name (via mfapi.in/AMFI), since the NAV in the
    imported Excel/PDF is only as fresh as the statement date. This
    replaces the old behavior of deferring NAV refresh to a manual button
    click after import."""
    funds=[]; source=parsed.get("source","CAS")
    for folio in parsed.get("folios",[]):
        amc=folio.get("amc","")
        faid=folio.get("_account_id",""); fanm=folio.get("_account_name",""); frel=folio.get("_relation","")
        for s in folio.get("schemes",[]):
            nm=s.get("scheme",""); isin=s.get("isin",""); amfi=s.get("amfi","")
            units=float(s.get("units",0) or 0); nav=float(s.get("nav",0) or 0)
            value=float(s.get("value",0) or 0); cost=float(s.get("cost",0) or 0)
            if not nm or (units==0 and value==0 and cost==0): continue

            # ── Resolve AMFI scheme code: direct code → ISIN search → name search ──
            sc=None
            if amfi and str(amfi).isdigit(): sc=int(amfi)
            if not sc and isin:
                try:
                    res = search_funds(isin)
                    if res: sc = res[0]["schemeCode"]
                except Exception:
                    pass
            if not sc and nm:
                try:
                    res = search_funds(nm[:30])
                    if res: sc = res[0]["schemeCode"]
                except Exception:
                    pass

            # ── Fetch live NAV now (falls back to the CAS/Excel NAV if the lookup fails) ──
            lnav, ldate = nav, s.get("valuation_date","")
            if sc:
                try:
                    nd = fetch_latest_nav(sc)
                    if nd:
                        lnav, ldate = nd["nav"], nd["date"]
                except Exception:
                    pass

            lval = units*lnav if units>0 and lnav>0 else (value if value>0 else 0.0)
            inv  = cost if cost>0 else (value if value>0 else round(units*nav,2))
            pnl  = round(lval - inv, 2)
            aid  = s.get("_account_id") or faid
            obj={
                "name":nm[:80],"scheme_code":sc,"isin":isin,"amfi":amfi,
                "amc":s.get("amc","") or amc,"folio":s.get("folio","") or folio.get("folio",""),
                "category":"Other","units":units,"avg_nav":round(inv/units,4) if units>0 else nav,
                "invested":inv,"current_value":lval,"latest_nav":lnav,"nav_date":ldate,
                "pnl":pnl,"cas_nav":nav,"cas_value":value,"sip":0,"since":"","since_date":"","source":source,
            }
            if aid:
                obj["_import_account_id"]   = aid
                obj["_import_account_name"] = s.get("_account_name",fanm)
                obj["_import_relation"]     = s.get("_relation",frel)
            funds.append(obj)
    return funds

def refresh_navs(funds: list) -> list:
    for f in funds:
        if f.get("scheme_code"):
            nd = fetch_latest_nav(f["scheme_code"])
            if nd:
                f["latest_nav"] = nd["nav"]
                f["nav_date"]   = nd["date"]
                if f.get("units", 0) > 0:
                    f["current_value"] = round(f["units"] * nd["nav"], 2)
                    f["pnl"] = round(f["current_value"] - f.get("invested", 0), 2)
    return funds


def sync_to_finance(store: dict):
    funds = all_funds(store)
    fa_list = [{
        "name":        f.get("name",""),
        "invested":    f.get("invested", 0),
        "value":       f.get("current_value", f.get("invested", 0)),
        "sip":         f.get("sip", 0),
        "scheme_code": f.get("scheme_code"),
        "units":       f.get("units", 0),
        "nav":         f.get("latest_nav", 0),
        "nav_date":    f.get("nav_date", ""),
    } for f in funds]
    fa_p = st.session_state.get("fa_profile", {})
    fa_p["mf_list"] = fa_list
    st.session_state.fa_profile = fa_p
    try:
        from finance_advisor import fa_save
        fa_save()
    except Exception:
        pass


# ══════════════════════════════════════════════════════════════
# CSS
# ══════════════════════════════════════════════════════════════
_CSS = """<style>
.mf-card{background:var(--bg-card);border:1px solid var(--border);border-radius:10px;padding:16px 20px;margin-bottom:10px;}
.mf-hdr{font-family:'DM Mono',monospace;font-size:9px;letter-spacing:1.5px;color:var(--text-muted);text-transform:uppercase;margin-bottom:8px;}
.mf-row{background:var(--bg-secondary);border:1px solid var(--border);border-radius:8px;padding:12px 14px;margin-bottom:6px;}
.mf-tag{display:inline-block;font-size:9px;padding:2px 8px;border-radius:10px;font-family:'DM Mono',monospace;font-weight:600;}
.mf-acc-badge{background:var(--bg-card);border:1px solid var(--accent-blue);border-radius:6px;padding:3px 8px;font-size:10px;color:var(--accent-blue);font-family:'DM Mono',monospace;}
</style>"""

def metric_card(label, value, color="var(--text-primary)", sub=""):
    sub_html = f'<div style="font-size:10px;color:var(--text-muted);">{sub}</div>' if sub else ""
    return (f'<div class="mf-card" style="text-align:center;padding:12px;">'
            f'<div class="mf-hdr">{label}</div>'
            f'<div style="font-size:16px;font-weight:700;color:{color};">{value}</div>'
            f'{sub_html}</div>')

def fund_row_html(f, show_account=False):
    inv  = f.get("invested", 0)
    cur  = f.get("current_value", inv)
    gain = cur - inv
    ar   = abs_ret(inv, cur)
    gc   = color_ret(ar)
    nav_s  = f"NAV Rs{f['latest_nav']:.4f}" if f.get("latest_nav") else ""
    date_s = f.get("nav_date", "")
    units_s= f"{f['units']:.3f} units" if f.get("units") else ""
    linked = "🔗" if f.get("scheme_code") else "📝"
    acc_badge = ""
    if show_account and f.get("_account_name"):
        acc_badge = (f'<span class="mf-acc-badge" style="margin-left:8px;">'
                     f'{f.get("_relation","")} · {f["_account_name"]}</span>')
    return (
        f'<div class="mf-row">'
        f'<div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:5px;">'
        f'<div><span style="font-size:12px;font-weight:700;color:var(--text-primary);">'
        f'{linked} {f.get("name","")[:52]}</span>{acc_badge}</div>'
        f'<div style="text-align:right;">'
        f'<div style="font-size:13px;font-weight:700;color:{gc};">{ar:+.2f}%</div>'
        f'<div style="font-size:10px;color:var(--text-muted);">{fmt(gain)}</div></div></div>'
        f'<div style="display:flex;gap:16px;flex-wrap:wrap;font-size:11px;">'
        f'<span style="color:var(--text-muted);">Inv: <b style="color:var(--text-primary);">{fmt(inv)}</b></span>'
        f'<span style="color:var(--text-muted);">Cur: <b style="color:var(--accent-green);">{fmt(cur)}</b></span>'
        f'{"<span style=\"color:var(--text-muted)\">" + units_s + "</span>" if units_s else ""}'
        f'{"<span style=\"font-size:10px;color:var(--text-muted)\">" + nav_s + " " + date_s + "</span>" if nav_s else ""}'
        f'</div></div>'
    )


# ══════════════════════════════════════════════════════════════
# MAIN PAGE
# ══════════════════════════════════════════════════════════════

def page_mf_portfolio():
    st.markdown(_CSS, unsafe_allow_html=True)
    st.markdown("""<div style="padding:8px 0 14px;">
      <div style="font-family:'Fraunces',serif;font-size:28px;font-weight:600;color:var(--text-primary);">
        💼 MF Portfolio Tracker</div>
      <div style="font-family:'DM Mono',monospace;font-size:10px;color:var(--text-muted);margin-top:2px;">
        MULTI-ACCOUNT · LIVE NAV · CDSL/NSDL CAS IMPORT · FAMILY PORTFOLIO</div>
    </div>""", unsafe_allow_html=True)

    store    = get_store()
    accounts = get_accounts(store)
    funds    = all_funds(store)

    t1, t2, t3, t4, t5 = st.tabs(
        ["📊 Overview", "👥 Accounts", "📂 Import CAS", "➕ Add Fund", "🔬 Analyze"]
    )

    # ═════════════════════════════ OVERVIEW ═══════════════════
    with t1:
        if not funds:
            st.markdown("""<div class="mf-card" style="text-align:center;padding:40px;">
              <div style="font-size:36px;margin-bottom:12px;">📭</div>
              <div style="font-size:15px;font-weight:600;color:var(--text-primary);">No funds yet</div>
              <div style="font-size:12px;color:var(--text-muted);margin-top:6px;">
                Import your CDSL monthly CAS PDF in <b>Import CAS</b>, or add funds in <b>Add Fund</b>.
              </div></div>""", unsafe_allow_html=True)
        else:
            rh1, rh2 = st.columns([4, 1])
            with rh2:
                if st.button("🔄 Refresh NAVs", key="refresh_navs", width='stretch'):
                    with st.spinner("Fetching live NAVs for all funds..."):
                        updated = 0
                        for acc_id, acc in store.get("accounts", {}).items():
                            for f in acc.get("funds", []):
                                isin = f.get("isin","")
                                nm   = f.get("name","")
                                sc   = f.get("scheme_code")
                                # Resolve scheme_code if missing
                                if not sc and isin:
                                    res = search_funds(isin)
                                    if res: sc = res[0]["schemeCode"]
                                if not sc and nm:
                                    res = search_funds(nm[:30])
                                    if res: sc = res[0]["schemeCode"]
                                if sc:
                                    f["scheme_code"] = sc
                                    nd = fetch_latest_nav(sc)
                                    if nd:
                                        f["latest_nav"]    = nd["nav"]
                                        f["nav_date"]      = nd["date"]
                                        f["current_value"] = round(f.get("units",0) * nd["nav"], 2)
                                        f["pnl"]           = round(f["current_value"] - f.get("invested", 0), 2)
                                        updated += 1
                        store["last_refresh"] = datetime.datetime.now().strftime("%d %b %Y %H:%M")
                        save_store(store)
                    st.success(f"✅ Updated NAVs for {updated} fund(s)")
                    st.rerun()
            with rh1:
                st.markdown(
                    f'<div style="font-size:11px;color:var(--text-muted);padding-top:6px;">'
                    f'Last refreshed: <b>{store.get("last_refresh","never")}</b> · '
                    f'<b>{len(funds)}</b> funds · <b>{len(accounts)}</b> account(s)</div>',
                    unsafe_allow_html=True)
            with rh2:
                if st.button("🔄 Refresh NAVs", key="ov_refresh", width='stretch'):
                    for acc in accounts.values():
                        acc["funds"] = refresh_navs(acc.get("funds", []))
                    store["funds"]        = refresh_navs(store.get("funds", []))
                    store["last_refresh"] = datetime.datetime.now().strftime("%d %b %Y %H:%M")
                    save_store(store)
                    sync_to_finance(store)
                    st.success("NAVs refreshed & synced to Finance Advisor!")
                    st.rerun()

            total_inv  = sum(f.get("invested", 0) for f in funds)
            total_cur  = sum(f.get("current_value", 0) for f in funds)
            total_gain = total_cur - total_inv
            total_sip  = sum(f.get("sip", 0) for f in funds)
            ar_all     = abs_ret(total_inv, total_cur)
            gc_all     = color_ret(ar_all)

            c1, c2, c3, c4, c5 = st.columns(5)
            for col, lbl, val, clr in [
                (c1, "INVESTED",    fmt(total_inv),   "var(--accent-blue)"),
                (c2, "CURRENT",     fmt(total_cur),   "var(--accent-green)"),
                (c3, "GAIN / LOSS", fmt(total_gain),  gc_all),
                (c4, "RETURNS",     f"{ar_all:+.2f}%", gc_all),
                (c5, "MONTHLY SIP", fmt(total_sip),   "var(--accent-gold)"),
            ]:
                with col:
                    st.markdown(metric_card(lbl, val, clr), unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)
            fl, fc = st.columns([3, 2])
            multi = len(accounts) > 1

            with fl:
                if multi:
                    st.markdown('<div class="mf-hdr">ALL ACCOUNTS — COMBINED</div>',
                                unsafe_allow_html=True)
                for fi, f in enumerate(funds):
                    acc_id    = f.get("_account_id", "manual")
                    fund_name = f.get("name", "")
                    ekey      = f"edit_{acc_id}_{fi}"

                    # ── Fund row + buttons ──────────────────────────
                    c_row, c_edit, c_del = st.columns([8, 1, 1])
                    with c_row:
                        st.markdown(fund_row_html(f, show_account=multi), unsafe_allow_html=True)
                    with c_edit:
                        if st.button("✏️", key=f"btn_{ekey}", help="Edit fund"):
                            cur = st.session_state.get(f"show_{ekey}", False)
                            st.session_state[f"show_{ekey}"] = not cur
                    with c_del:
                        if st.button("✗", key=f"del_{acc_id}_{fi}_{fund_name[:10]}", help="Delete"):
                            if acc_id == "manual":
                                store["funds"] = [x for x in store.get("funds", [])
                                                  if x.get("name") != fund_name]
                            elif acc_id in accounts:
                                accounts[acc_id]["funds"] = [
                                    x for x in accounts[acc_id].get("funds", [])
                                    if x.get("name") != fund_name
                                ]
                            save_store(store)
                            sync_to_finance(store)
                            st.rerun()

                    # ── Inline edit form ────────────────────────────
                    if st.session_state.get(f"show_{ekey}", False):
                        with st.container():
                            st.markdown(
                                f'<div style="background:var(--bg-secondary);border:1px solid var(--accent-blue);'
                                f'border-radius:8px;padding:14px 16px;margin:-4px 0 10px 0;">'
                                f'<b style="font-size:12px;">Edit — {fund_name[:55]}</b></div>',
                                unsafe_allow_html=True
                            )
                            ea, eb, ec = st.columns(3)
                            with ea:
                                e_name  = st.text_input("Fund Name",     value=f.get("name",""),           key=f"{ekey}_name")
                                e_inv   = st.number_input("Invested (Rs)", value=float(f.get("invested",0)), min_value=0.0, step=100.0, key=f"{ekey}_inv")
                                e_cur   = st.number_input("Current Value (Rs)", value=float(f.get("current_value",0)), min_value=0.0, step=100.0, key=f"{ekey}_cur")
                            with eb:
                                e_units = st.number_input("Units",        value=float(f.get("units",0)),    min_value=0.0, step=0.001, format="%.3f", key=f"{ekey}_units")
                                e_nav   = st.number_input("Latest NAV (Rs)", value=float(f.get("latest_nav",0)), min_value=0.0, step=0.01, key=f"{ekey}_nav")
                                e_avg   = st.number_input("Avg Buy NAV (Rs)", value=float(f.get("avg_nav",0)), min_value=0.0, step=0.01, key=f"{ekey}_avg")
                            with ec:
                                e_sip   = st.number_input("Monthly SIP (Rs)", value=float(f.get("sip",0)), min_value=0.0, step=100.0, key=f"{ekey}_sip")
                                e_since = st.text_input("Invested Since", value=f.get("since",""),          key=f"{ekey}_since")
                                e_cat   = st.selectbox("Category", MF_CATEGORIES,
                                            index=MF_CATEGORIES.index(f.get("category","Other")) if f.get("category","Other") in MF_CATEGORIES else len(MF_CATEGORIES)-1,
                                            key=f"{ekey}_cat")

                            # Auto-calc current value if NAV+units filled and cur=0
                            if e_nav > 0 and e_units > 0 and e_cur == 0:
                                e_cur = round(e_nav * e_units, 2)

                            sv1, sv2, _ = st.columns([1, 1, 4])
                            with sv1:
                                if st.button("💾 Save", key=f"{ekey}_save", type="primary"):
                                    updated = {
                                        "name":          e_name.strip() or fund_name,
                                        "invested":      e_inv,
                                        "current_value": e_cur if e_cur > 0 else round(e_nav * e_units, 2),
                                        "units":         e_units,
                                        "latest_nav":    e_nav,
                                        "avg_nav":       e_avg,
                                        "sip":           e_sip,
                                        "since":         e_since,
                                        "category":      e_cat,
                                    }
                                    # Write back into store
                                    if acc_id == "manual":
                                        for x in store.get("funds", []):
                                            if x.get("name") == fund_name:
                                                x.update(updated)
                                    elif acc_id in accounts:
                                        for x in accounts[acc_id].get("funds", []):
                                            if x.get("name") == fund_name:
                                                x.update(updated)
                                        store["accounts"] = accounts
                                    save_store(store)
                                    sync_to_finance(store)
                                    st.session_state[f"show_{ekey}"] = False
                                    st.rerun()
                            with sv2:
                                if st.button("✕ Cancel", key=f"{ekey}_cancel"):
                                    st.session_state[f"show_{ekey}"] = False
                                    st.rerun()

            with fc:
                _donut(funds)

            # ── SIP Manager ──────────────────────────────────────
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown('<div class="mf-hdr">MONTHLY SIP MANAGER — SET YOUR SIP PER FUND</div>',
                        unsafe_allow_html=True)
            sip_funds_exist = any(f.get("sip",0) > 0 for f in funds)
            if not sip_funds_exist:
                st.caption("💡 Set your monthly SIP amount per fund below — these will reflect in Finance Advisor budget.")

            sip_cols = st.columns(2)
            for si, f in enumerate(funds):
                acc_id    = f.get("_account_id","manual")
                fund_name = f.get("name","")
                cur_sip   = float(f.get("sip", 0))
                with sip_cols[si % 2]:
                    sc1, sc2, sc3 = st.columns([4, 2, 1])
                    with sc1:
                        st.markdown(
                            f'<div style="font-size:11px;font-weight:600;color:var(--text-primary);'
                            f'padding-top:8px;overflow:hidden;white-space:nowrap;text-overflow:ellipsis;">'
                            f'{fund_name[:38]}</div>',
                            unsafe_allow_html=True
                        )
                    with sc2:
                        new_sip = st.number_input(
                            "SIP ₹/mo", value=cur_sip, min_value=0.0, step=500.0,
                            key=f"sip_quick_{acc_id}_{si}", label_visibility="collapsed"
                        )
                    with sc3:
                        if st.button("✓", key=f"sip_save_{acc_id}_{si}", help="Save SIP"):
                            if acc_id == "manual":
                                for x in store.get("funds", []):
                                    if x.get("name") == fund_name:
                                        x["sip"] = new_sip
                            elif acc_id in accounts:
                                for x in accounts[acc_id].get("funds", []):
                                    if x.get("name") == fund_name:
                                        x["sip"] = new_sip
                                store["accounts"] = accounts
                            save_store(store)
                            sync_to_finance(store)
                            st.rerun()

            # Show total
            total_sip_now = sum(f.get("sip",0) for f in all_funds(store))
            if total_sip_now > 0:
                st.markdown(
                    f'<div style="background:var(--bg-secondary);border:1px solid var(--accent-gold);'
                    f'border-radius:8px;padding:10px 16px;margin-top:8px;display:flex;justify-content:space-between;">'
                    f'<span style="font-size:12px;color:var(--text-muted);">TOTAL MONTHLY SIP</span>'
                    f'<span style="font-size:16px;font-weight:700;color:var(--accent-gold);">{fmt(total_sip_now)}</span>'
                    f'</div>',
                    unsafe_allow_html=True
                )
                st.caption("✅ This total is synced to Finance Advisor → Budget → Committed Outflows")

            # Per-account breakdown
            if multi:
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown('<div class="mf-hdr">PER-ACCOUNT BREAKDOWN</div>', unsafe_allow_html=True)
                acc_cols = st.columns(min(len(accounts), 3))
                for i, (acc_id, acc) in enumerate(accounts.items()):
                    af  = acc.get("funds", [])
                    ai  = sum(f.get("invested", 0) for f in af)
                    ac2 = sum(f.get("current_value", 0) for f in af)
                    ag  = ac2 - ai
                    ar2 = abs_ret(ai, ac2)
                    gc2 = color_ret(ar2)
                    with acc_cols[i % 3]:
                        st.markdown(f"""<div class="mf-card">
                          <div class="mf-hdr">{acc.get("relation","")}</div>
                          <div style="font-size:14px;font-weight:700;color:var(--text-primary);">{acc.get("name","")}</div>
                          <div style="font-size:11px;color:var(--text-muted);margin-top:3px;">
                            PAN: {mask_pan(acc.get("pan"))} · {len(af)} funds</div>
                          <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:10px;">
                            <div><div class="mf-hdr">INVESTED</div>
                              <div style="font-size:13px;font-weight:600;color:var(--accent-blue);">{fmt(ai)}</div></div>
                            <div><div class="mf-hdr">CURRENT</div>
                              <div style="font-size:13px;font-weight:600;color:var(--accent-green);">{fmt(ac2)}</div></div>
                            <div><div class="mf-hdr">GAIN</div>
                              <div style="font-size:13px;font-weight:600;color:{gc2};">{fmt(ag)}</div></div>
                            <div><div class="mf-hdr">RETURNS</div>
                              <div style="font-size:13px;font-weight:600;color:{gc2};">{ar2:+.2f}%</div></div>
                          </div></div>""", unsafe_allow_html=True)

    # ═════════════════════════════ ACCOUNTS ═══════════════════
    with t2:
        st.markdown('<div class="mf-hdr" style="margin-bottom:12px;">FAMILY ACCOUNTS</div>',
                    unsafe_allow_html=True)

        if accounts:
            for acc_id, acc in list(accounts.items()):
                af  = acc.get("funds", [])
                ai  = sum(f.get("invested", 0) for f in af)
                ac2 = sum(f.get("current_value", 0) for f in af)
                ca1, ca2, ca3 = st.columns([4, 2, 1])
                with ca1:
                    src  = acc.get("source", "Manual")
                    stmt = acc.get("statement_period", "")
                    st.markdown(f"""<div class="mf-row">
                      <div style="display:flex;align-items:center;gap:10px;margin-bottom:4px;">
                        <span style="font-size:13px;font-weight:700;color:var(--text-primary);">
                          {acc.get("relation","")} — {acc.get("name","")}</span>
                        <span class="mf-tag" style="background:var(--bg-card);
                          color:var(--accent-blue);border:1px solid var(--border);">{src}</span>
                      </div>
                      <div style="font-size:11px;color:var(--text-muted);">
                        PAN: {mask_pan(acc.get("pan"))} · {len(af)} funds · Inv: {fmt(ai)} · Cur: {fmt(ac2)}
                        {f" · Period: {stmt}" if stmt else ""}
                      </div></div>""", unsafe_allow_html=True)
                with ca2:
                    st.markdown(f'<div style="padding:8px 0;font-size:13px;font-weight:700;'
                                f'color:var(--accent-green);">{fmt(ac2)}</div>',
                                unsafe_allow_html=True)
                with ca3:
                    if st.button("🗑", key=f"del_acc_{acc_id}"):
                        st.session_state[f"confirm_del_{acc_id}"] = True
                if st.session_state.get(f"confirm_del_{acc_id}"):
                    st.warning(f"Delete **{acc.get('name','')}** and all {len(af)} funds?")
                    dc1, dc2, _ = st.columns([1,1,4])
                    with dc1:
                        if st.button("Yes, delete", key=f"yes_{acc_id}", type="primary"):
                            del accounts[acc_id]
                            st.session_state.pop(f"confirm_del_{acc_id}", None)
                            save_store(store)
                            sync_to_finance(store)
                            st.rerun()
                    with dc2:
                        if st.button("Cancel", key=f"no_{acc_id}"):
                            st.session_state.pop(f"confirm_del_{acc_id}", None)
                            st.rerun()
        else:
            st.info("No accounts yet. Import a CAS PDF to create one automatically.")

        st.markdown("---")
        st.markdown('<div class="mf-hdr" style="margin-bottom:10px;">ADD ACCOUNT MANUALLY</div>',
                    unsafe_allow_html=True)
        am1, am2, am3 = st.columns(3)
        with am1: acc_name_i = st.text_input("Name", key="acc_name_i", placeholder="Priya Kumar")
        with am2: acc_rel_i  = st.selectbox("Relation", ACCOUNT_RELATIONS, key="acc_rel_i")
        with am3: acc_pan_i  = st.text_input("PAN (optional)", key="acc_pan_i")
        if st.button("Add Account", key="add_acc_btn", type="primary"):
            if acc_name_i.strip():
                aid = f"acc_{len(accounts)+1}_{int(datetime.datetime.now().timestamp())}"
                accounts[aid] = {
                    "name": acc_name_i.strip(), "relation": acc_rel_i,
                    "pan": acc_pan_i.strip().upper(), "funds": [],
                    "source": "Manual", "statement_period": "",
                    "import_date": datetime.date.today().isoformat(),
                }
                save_store(store)
                st.success(f"Account **{acc_name_i}** created. Import CAS or add funds manually.")
                st.rerun()
            else:
                st.error("Enter the account holder name.")


    # ═════════════════════════════ IMPORT CAS ═════════════════
    with t3:
        # ── Show post-import confirmation (survives rerun) ──────
        if st.session_state.get("_import_ok"):
            msg = st.session_state.pop("_import_ok")
            st.success(msg)

        # ── Step 1: Upload & Parse ───────────────────────────────
        st.markdown("### Step 1 — Upload your CAS statement")
        st.markdown('<div style="font-size:11px;color:var(--text-muted);margin-bottom:8px;">Your real CDSL/NSDL Consolidated Account Statement is a <b>PDF</b> — that\'s what you\'ll usually want. Excel is also supported if you\'ve exported your holdings that way instead.</div>', unsafe_allow_html=True)

        cas_format = st.radio("Format", ["📄 PDF (CDSL/NSDL CAS)", "📊 Excel"],
                               horizontal=True, key="cas_format_choice", label_visibility="collapsed")

        if cas_format.startswith("📄"):
            cas_file = st.file_uploader(
                "Upload CAS PDF", type=["pdf"], key="cas_pdf_up",
                help="The monthly/annual CAS PDF emailed by CDSL (cdslindia.com) or NSDL — usually password-protected."
            )
            cas_password = st.text_input(
                "PDF Password (if protected)", type="password", key="cas_pdf_pw",
                help="Usually your PAN in uppercase, or PAN+DOB depending on the issuer — check the email that sent you the CAS."
            )
            if cas_file:
                file_bytes = cas_file.read()
                st.info(f"📄 **{cas_file.name}** ready")
                if st.button("📥 Parse PDF", key="parse_pdf_btn", type="primary"):
                    with st.spinner("Reading CAS PDF — this can take a few seconds…"):
                        parsed = parse_cas_pdf(file_bytes, cas_password)
                    if not parsed.get("ok"):
                        st.error("Parse failed: " + " | ".join(parsed.get("errors", ["Unknown error"])))
                        st.markdown('<div style="font-size:11px;color:var(--text-muted);">Common fixes: double-check the password, or make sure this is a genuine CDSL/NSDL CAS PDF (not a screenshot/scan).</div>', unsafe_allow_html=True)
                    else:
                        st.session_state["_cas_parsed"] = parsed
                        st.session_state["_cas_is_excel"] = False
                        nf = len(parsed.get("folios", []))
                        ns = sum(len(x["schemes"]) for x in parsed.get("folios", []))
                        st.success(f"✅ Parsed: {nf} folio(s), {ns} fund(s) found")
        else:
            cas_file = st.file_uploader(
                "Upload MF Portfolio Excel (.xlsx)",
                type=["xlsx","xls"], key="cas_excel_up"
            )
            if cas_file:
                file_bytes = cas_file.read()
                st.info(f"📊 **{cas_file.name}** ready")
                if st.button("📥 Parse Excel", key="parse_btn", type="primary"):
                    with st.spinner("Reading Excel..."):
                        parsed = parse_mf_excel(file_bytes)
                    if not parsed.get("ok"):
                        st.error("Parse failed: " + " | ".join(parsed.get("errors",[])))
                    else:
                        st.session_state["_cas_parsed"] = parsed
                        st.session_state["_cas_is_excel"] = True
                        nf = len(parsed.get("folios",[]))
                        ns = sum(len(x["schemes"]) for x in parsed.get("folios",[]))
                        st.success(f"✅ Parsed: {nf} folio(s), {ns} fund(s) found")

        # ── Step 2: Import ───────────────────────────────────────
        if st.session_state.get("_cas_parsed"):
            pc = st.session_state["_cas_parsed"]
            ns = sum(len(x["schemes"]) for x in pc.get("folios",[]))
            inv = pc.get("investor",{})

            st.markdown(f"### Step 2 — Import {ns} Funds")
            st.write(f"Investor: **{inv.get('name','')}** | Source: `{pc.get('source')}`")

            # Preview funds
            with st.expander(f"Preview {ns} funds", expanded=False):
                for folio in pc.get("folios",[]):
                    for s in folio.get("schemes",[]):
                        st.write(f"• {s.get('scheme','')} — {s.get('units',0)} units — Rs {s.get('cost',0):,.0f} invested")

            col1, col2 = st.columns([1,4])
            with col1:
                if st.button("🚀 Import Funds", key="do_import", type="primary", width='stretch'):
                    try:
                        with st.spinner("Importing..."):
                            new_funds = cas_to_funds(pc)

                        if not new_funds:
                            st.error(f"cas_to_funds returned 0 funds!")
                        else:
                            inv_info   = pc.get("investor", {})
                            per        = pc.get("statement_period", {})
                            period_str = f"{per.get('from','')} to {per.get('to','')}"
                            total_added = 0

                            # Group by account
                            by_acc = {}
                            for f in new_funds:
                                by_acc.setdefault(f.get("_import_account_id","SELF"), []).append(f)

                            for imp_id, flist in by_acc.items():
                                ff       = flist[0]
                                imp_name = ff.get("_import_account_name", imp_id)
                                imp_rel  = ff.get("_import_relation", "Self")

                                # Find or create account
                                mid = None
                                for eid, ea in accounts.items():
                                    if ea.get("name","").lower() == imp_name.lower():
                                        mid = eid; break
                                if not mid:
                                    import time
                                    mid = f"mf_{imp_id}_{int(time.time())}"
                                    accounts[mid] = {
                                        "name":             imp_name,
                                        "relation":         imp_rel,
                                        "pan":              inv_info.get("pan",""),
                                        "email":            inv_info.get("email",""),
                                        "mobile":           inv_info.get("mobile",""),
                                        "source":           "Excel Template",
                                        "statement_period": period_str,
                                        "import_date":      str(__import__("datetime").date.today()),
                                        "funds":            [],
                                    }

                                # Add funds
                                clean    = [{k:v for k,v in f.items() if not k.startswith("_import_")} for f in flist]
                                existing = accounts[mid]["funds"]
                                ex_names = {f["name"].lower() for f in existing}
                                for f in clean:
                                    if f["name"].lower() not in ex_names:
                                        existing.append(f)
                                        total_added += 1
                                accounts[mid]["funds"] = existing

                            # Save
                            store["accounts"] = accounts
                            store["last_refresh"] = str(__import__("datetime").datetime.now().strftime("%d %b %Y %H:%M"))
                            st.session_state.mf_store = store

                            # Write to disk
                            import json, os
                            _path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "equitex_mf_portfolio.json")
                            with open(_path,"w") as _jf:
                                json.dump(store, _jf, indent=2, default=str)

                            sync_to_finance(store)
                            st.session_state.pop("_cas_parsed", None)
                            st.session_state.pop("_cas_is_excel", None)
                            st.session_state["_import_ok"] = f"✅ {total_added} fund(s) imported for {list(accounts.values())[-1].get('name','')} — live NAVs fetched automatically (not the statement's old NAV)."
                            st.rerun()

                    except Exception as _e:
                        import traceback
                        st.error(f"❌ Error: {_e}")
                        st.code(traceback.format_exc())

            with col2:
                if st.button("✕ Discard", key="discard_cas"):
                    st.session_state.pop("_cas_parsed", None)
                    st.session_state.pop("_cas_is_excel", None)
                    st.rerun()

        else:
            st.markdown("*Upload and parse an Excel file above to get started.*")


    with t4:
        st.markdown('<div class="mf-hdr" style="margin-bottom:12px;">SEARCH FUND & ADD WITH LIVE NAV</div>',
                    unsafe_allow_html=True)

        if accounts:
            add_opts = {"No account (manual pool)": "manual"}
            for aid, acc in accounts.items():
                add_opts[f"{acc.get('relation','')} — {acc.get('name','')}"] = aid
            add_lbl = st.selectbox("Add to:", list(add_opts.keys()), key="add_acc_sel")
            add_aid = add_opts[add_lbl]
        else:
            add_aid = "manual"
            st.info("No accounts yet — fund will go into the manual pool.")

        sq = st.text_input("Search fund name",
                           placeholder="e.g. Mirae Asset Large Cap, HDFC Mid Cap",
                           key="af_search")
        sel_fund = None
        if sq and len(sq) >= 3:
            with st.spinner("Searching AMFI..."):
                results = search_funds(sq)
            if results:
                opts = {f"{r['schemeName']} (#{r['schemeCode']})": r for r in results}
                chosen = st.selectbox("Select", list(opts.keys()), key="af_sel")
                sel_fund = opts[chosen]
                nd = fetch_latest_nav(sel_fund["schemeCode"])
                if nd:
                    st.markdown(f"""<div style="background:var(--bg-card);
                      border:1px solid var(--accent-green);border-radius:8px;
                      padding:9px 14px;margin:6px 0;font-size:12px;">
                      NAV: <b style="color:var(--accent-green);">Rs {nd['nav']:.4f}</b>
                      as of {nd['date']}
                    </div>""", unsafe_allow_html=True)
            else:
                st.warning("No matching funds found.")

        d1, d2, d3 = st.columns(3)
        with d1:
            mf_name = st.text_input("Fund name *",
                                     value=sel_fund["schemeName"] if sel_fund else "",
                                     key="af_name")
        with d2: mf_cat = st.selectbox("Category", MF_CATEGORIES, key="af_cat")
        with d3: mf_sip = st.number_input("Monthly SIP Rs", 0, 1000000, 0, 500, key="af_sip")

        d4, d5, d6 = st.columns(3)
        with d4: mf_units   = st.number_input("Units held", 0.0, 1e7, 0.0, 0.001,
                                               format="%.3f", key="af_units")
        with d5: mf_avg_nav = st.number_input("Avg purchase NAV Rs", 0.0, 1e5, 0.0, 0.01,
                                               format="%.4f", key="af_avg_nav")
        with d6: mf_inv     = st.number_input("Invested Rs (0=auto)", 0, 10_000_000, 0, 1000,
                                               key="af_inv")
        d7, d8 = st.columns(2)
        with d7: mf_since = st.text_input("Since MM/YYYY", placeholder="01/2021", key="af_since")
        with d8: mf_folio = st.text_input("Folio (optional)", key="af_folio")

        auto_inv = mf_units * mf_avg_nav if mf_units > 0 and mf_avg_nav > 0 else 0
        disp_inv = mf_inv if mf_inv > 0 else auto_inv
        if auto_inv > 0 and mf_inv == 0:
            st.caption(f"Auto-calc invested: {fmt(auto_inv)}")

        if st.button("Add Fund", key="do_add", type="primary"):
            if not mf_name.strip():
                st.error("Fund name required.")
            else:
                sc   = sel_fund["schemeCode"] if sel_fund else None
                nd   = fetch_latest_nav(sc) if sc else {}
                lnav = nd.get("nav", mf_avg_nav or 0)
                ldt  = nd.get("date", "")
                cur  = mf_units * lnav if mf_units > 0 and lnav > 0 else disp_inv
                sd   = ""
                try:
                    m2, y2 = mf_since.split("/")
                    sd = datetime.date(int(y2), int(m2), 1).isoformat()
                except Exception:
                    pass

                nf = {
                    "name": mf_name.strip(), "scheme_code": sc,
                    "isin": "", "amfi": "", "amc": "", "folio": mf_folio,
                    "category": mf_cat, "units": mf_units, "avg_nav": mf_avg_nav,
                    "invested": disp_inv, "current_value": cur,
                    "latest_nav": lnav, "nav_date": ldt,
                    "sip": mf_sip, "since": mf_since, "since_date": sd,
                    "source": "Manual",
                }
                if add_aid == "manual":
                    store.setdefault("funds", []).append(nf)
                else:
                    accounts[add_aid].setdefault("funds", []).append(nf)

                store["last_refresh"] = datetime.datetime.now().strftime("%d %b %Y %H:%M")
                save_store(store)
                sync_to_finance(store)
                for k in ["af_search","af_name","af_units","af_avg_nav",
                          "af_inv","af_sip","af_since","af_folio","af_sel"]:
                    st.session_state.pop(k, None)
                st.success(f"Added **{mf_name.strip()}** — current value {fmt(cur)}")
                st.rerun()

    # ═════════════════════════════ ANALYZE ════════════════════
    with t5:
        if not funds:
            st.info("Add funds first to see analysis.")
            return

        if accounts:
            af_opts = {"All accounts": None}
            for aid, acc in accounts.items():
                af_opts[f"{acc.get('relation','')} — {acc.get('name','')}"] = aid
            af_lbl = st.selectbox("Filter by account", list(af_opts.keys()), key="an_filter")
            af_id  = af_opts[af_lbl]
            filtered = [f for f in funds if af_id is None or f.get("_account_id") == af_id]
        else:
            filtered = funds

        if not filtered:
            st.info("No funds in selected account.")
            return

        idx = st.selectbox(
            "Select fund",
            range(len(filtered)),
            format_func=lambda i: (
                f"{filtered[i].get('name','')[:50]}"
                + (f" — {filtered[i].get('_account_name','')}" if len(accounts) > 1 else "")
            ),
            key="an_sel"
        )
        f  = filtered[idx]
        sc = f.get("scheme_code")

        history = []
        if sc:
            with st.spinner("Loading NAV history..."):
                history = fetch_nav_history(sc, 365)

        inv  = f.get("invested", 0)
        cur  = f.get("current_value", inv)
        units= f.get("units", 0)
        sip  = f.get("sip", 0)
        ar   = abs_ret(inv, cur)
        xi   = xirr_est(inv, cur, sip, f.get("since_date",""))
        gc   = color_ret(ar)

        st.markdown(f"""<div class="mf-card">
          <div style="font-size:15px;font-weight:700;color:var(--text-primary);margin-bottom:12px;">
            {f.get('name','')}
            {f'<span class="mf-tag" style="background:var(--bg-secondary);color:var(--accent-blue);border:1px solid var(--border);margin-left:10px;">{f.get("category","")}</span>' if f.get("category") else ''}
          </div>
          <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:12px;">
            <div><div class="mf-hdr">INVESTED</div>
              <div style="font-size:14px;font-weight:700;color:var(--accent-blue);">{fmt(inv)}</div></div>
            <div><div class="mf-hdr">CURRENT</div>
              <div style="font-size:14px;font-weight:700;color:var(--accent-green);">{fmt(cur)}</div></div>
            <div><div class="mf-hdr">GAIN / LOSS</div>
              <div style="font-size:14px;font-weight:700;color:{gc};">{fmt(cur-inv)}</div></div>
            <div><div class="mf-hdr">ABS RETURN</div>
              <div style="font-size:14px;font-weight:700;color:{gc};">{ar:+.2f}%</div></div>
            <div><div class="mf-hdr">UNITS</div>
              <div style="font-size:14px;font-weight:700;">{units:.3f}</div></div>
            <div><div class="mf-hdr">LATEST NAV</div>
              <div style="font-size:14px;font-weight:700;">Rs {f.get('latest_nav',0):.4f}</div></div>
            <div><div class="mf-hdr">AVG COST NAV</div>
              <div style="font-size:14px;font-weight:700;">Rs {f.get('avg_nav',0):.4f}</div></div>
            <div><div class="mf-hdr">EST. XIRR</div>
              <div style="font-size:14px;font-weight:700;color:{color_ret(xi)};">{xi:+.2f}%</div></div>
          </div></div>""", unsafe_allow_html=True)

        if history:
            st.markdown('<div class="mf-hdr" style="margin:12px 0 6px;">PERIOD RETURNS</div>',
                        unsafe_allow_html=True)
            pcols = st.columns(5)
            for col, (lbl, d) in zip(pcols, [("1W",7),("1M",30),("3M",90),("6M",180),("1Y",365)]):
                ret = nav_period_ret(history, min(d, len(history)))
                with col:
                    st.markdown(
                        metric_card(lbl, f"{ret:+.2f}%" if ret is not None else "N/A",
                                    color_ret(ret) if ret is not None else "var(--text-muted)"),
                        unsafe_allow_html=True
                    )

            try:
                import plotly.graph_objects as go
                dates, navs = [], []
                for h in reversed(history):
                    try:
                        dates.append(datetime.datetime.strptime(h["date"], "%d-%m-%Y"))
                        navs.append(float(h["nav"]))
                    except Exception:
                        pass
                if dates:
                    fig = go.Figure()
                    fig.add_trace(go.Scatter(x=dates, y=navs, mode="lines",
                        line=dict(color="#4a9eff", width=2),
                        fill="tozeroy", fillcolor="rgba(74,158,255,0.07)"))
                    if f.get("avg_nav"):
                        fig.add_hline(y=f["avg_nav"], line_dash="dash", line_color="#c9a84c",
                            annotation_text=f"Avg Cost Rs{f['avg_nav']:.2f}",
                            annotation_position="bottom right")
                    fig.update_layout(
                        title=f"1-Year NAV — {f.get('name','')[:40]}",
                        height=260, margin=dict(l=10,r=10,t=40,b=10),
                        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                        font=dict(color="#a8b8cc", size=11),
                        xaxis=dict(gridcolor="rgba(100,120,150,0.15)"),
                        yaxis=dict(gridcolor="rgba(100,120,150,0.15)", tickprefix="Rs"),
                        showlegend=False,
                    )
                    st.plotly_chart(fig, width='stretch')
            except Exception:
                pass

        if sip > 0:
            st.markdown('<div class="mf-hdr" style="margin:12px 0 6px;">SIP PROJECTION</div>',
                        unsafe_allow_html=True)
            annual = max(6.0, xi) / 100
            mr     = annual / 12
            sp2 = st.columns(5)
            for col, yr in zip(sp2, [1,3,5,10,15]):
                months = yr * 12
                fv = sip * ((1+mr)**months - 1) / mr * (1+mr) if mr > 0 else sip * months
                with col:
                    st.markdown(metric_card(f"{yr}Y", fmt(fv), "var(--accent-green)",
                                            sub=f"@{annual*100:.1f}%"),
                                unsafe_allow_html=True)

        _fund_signals(f, ar, xi, history)


# ── Donut chart helper ────────────────────────────────────────
def _donut(funds):
    try:
        import plotly.graph_objects as go
        cats = {}
        for f in funds:
            c = f.get("category","Other") or "Other"
            cats[c] = cats.get(c, 0) + f.get("current_value", f.get("invested", 0))
        if not cats: return
        labels = list(cats.keys())
        values = list(cats.values())
        total  = sum(values)
        colors = ["#4a9eff","#3ecf8e","#c9a84c","#a78bfa","#e05252",
                  "#2dd4bf","#f97316","#ec4899","#84cc16","#06b6d4"]
        fig = go.Figure(go.Pie(
            labels=labels, values=values, hole=0.55,
            marker=dict(colors=colors[:len(labels)],
                        line=dict(color="rgba(0,0,0,0.3)", width=1)),
            textinfo="label+percent",
            textfont=dict(size=10, color="#a8b8cc"),
        ))
        fig.update_layout(
            annotations=[dict(text=f"<b>{fmt(total)}</b>", x=0.5, y=0.5,
                              font=dict(size=13, color="#e2e8f0"), showarrow=False)],
            showlegend=False, height=260,
            margin=dict(l=5,r=5,t=5,b=5),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        )
        st.plotly_chart(fig, width='stretch')
    except Exception:
        pass


# ── Fund signals ──────────────────────────────────────────────
def _fund_signals(f, ar, xi, history):
    st.markdown('<div class="mf-hdr" style="margin:12px 0 6px;">QUICK ANALYSIS</div>',
                unsafe_allow_html=True)
    signals = []
    if xi >= 15:
        signals.append(("✅","Strong returns",f"XIRR {xi:.1f}% — above equity benchmarks","green"))
    elif xi >= 10:
        signals.append(("🟡","Decent returns",f"XIRR {xi:.1f}% — in line with long-term equity average","gold"))
    elif xi >= 0:
        signals.append(("⚠️","Below-average",f"XIRR {xi:.1f}% — review if this matches your goals","red"))
    else:
        signals.append(("🔴","Negative returns",f"XIRR {xi:.1f}% — consider exit if fundamentals haven't changed","red"))

    if len(history) >= 90:
        r1 = nav_period_ret(history, 30)
        r3 = nav_period_ret(history, 90)
        if r1 is not None and r3 is not None:
            if r1 > 0 and r3 > 0:
                signals.append(("✅","Positive momentum",f"1M: {r1:+.2f}% · 3M: {r3:+.2f}%","green"))
            elif r1 < 0 and r3 < 0:
                signals.append(("⚠️","Short-term weakness",
                                f"1M: {r1:+.2f}% · 3M: {r3:+.2f}% — could recover; check AMC news","gold"))

    if f.get("sip", 0) > 0:
        signals.append(("✅","Active SIP",f"Rs {f['sip']:,}/mo — rupee cost averaging in effect","green"))
    else:
        signals.append(("💡","No SIP linked","Consider starting a SIP for rupee cost averaging","blue"))

    cm = {"green":"var(--accent-green)","gold":"var(--accent-gold)",
          "red":"var(--accent-red)","blue":"var(--accent-blue)"}
    rows = "".join(
        f'<div style="display:flex;gap:12px;padding:9px 0;border-bottom:1px solid var(--border);">'
        f'<div style="font-size:16px;line-height:1;">{icon}</div>'
        f'<div><div style="font-size:12px;font-weight:600;color:{cm.get(c, "var(--text-primary)}")}">'
        f'{title}</div>'
        f'<div style="font-size:11px;color:var(--text-secondary);margin-top:1px;">{detail}</div>'
        f'</div></div>'
        for icon, title, detail, c in signals
    )
    st.markdown(f'<div class="mf-card">{rows}</div>', unsafe_allow_html=True)
    st.caption("XIRR is an estimate. Past returns don't guarantee future performance.")


# ── Backwards-compat exports for finance_advisor.py ──────────
def get_mf_portfolio() -> dict:
    store = get_store()
    return {**store, "funds": all_funds(store)}

def set_mf_portfolio(port: dict):
    store = get_store()
    if "funds" in port:
        store["funds"] = port["funds"]
    save_store(store)

def _refresh_all_navs(funds: list) -> list:
    return refresh_navs(funds)

def sync_mf_to_finance(port: dict):
    sync_to_finance(get_store())
